import requests
import openpyxl
import re

# Definir la URL base de la API de TMDb
base_url = "https://api.themoviedb.org/3/"

# Definir la clave de API de TMDb
api_key = ""

# Cargar el archivo de Excel que contiene la lista de películas
workbook = openpyxl.load_workbook("Listado de películas.xlsx")

worksheet = workbook.active

# Obtener los nombres de las películas desde la columna A, desde la fila 2 hasta la fila 100
nombres_peliculas = [worksheet.cell(row=i, column=3).value for i in range(1, 23555)]
# Obtener los tipos de películas desde la columna C, desde la fila 2 hasta la fila 100
tipos_peliculas = [worksheet.cell(row=i, column=4).value for i in range(1, 23555)]

# Agregar los tipos de películas desde el archivo de Excel a la lista
for row in worksheet.iter_rows(min_row=2, values_only=True):
    tipos_peliculas.append(row[2])

# Iterar sobre la lista de películas y descargar las portadas correspondientes
for index, pelicula in enumerate(nombres_peliculas):
    tipo_pelicula = tipos_peliculas[index]
    if tipo_pelicula == "MOVIE":
        # Definir la URL completa de la solicitud para películas
        url = f"{base_url}search/movie?api_key={api_key}&query={pelicula}"

        # Hacer la solicitud HTTP GET a la API de TMDb para películas
        response = requests.get(url)

        # Obtener la URL de la imagen de portada de la respuesta JSON
        results = response.json()["results"]
        if len(results) > 0:
            for result in results:
                if result["title"] == pelicula:
                    poster_path = result["poster_path"]
                    print(f"Portada de {pelicula} encontrada.")

                    # Construir la URL completa de la imagen de portada
                    poster_url = f"https://image.tmdb.org/t/p/w500/{poster_path}"

                    # Descargar la imagen de portada
                    poster_response = requests.get(poster_url)

                    # Guardar la imagen de portada en un archivo
                    file_name = re.sub(r'[<>:"/\\|?*\n]+', '_', pelicula).replace('.', '').replace(' ', '_').replace(':', '').replace('\'', '') + ".jpg"
                    with open(file_name, "wb") as f:
                        f.write(poster_response.content)
                    break
        else:
            print(f"No se encontró una portada para {pelicula}.")
    elif tipo_pelicula == "SHOW":
        # Definir la URL completa de la solicitud para series
        url = f"{base_url}search/tv?api_key={api_key}&query={pelicula}"

        # Hacer la solicitud HTTP GET a la API de TMDb para series
        response = requests.get(url)

        # Obtener la URL de la imagen de portada de la respuesta JSON
        results = response.json()["results"]
        if len(results) > 0:
            for result in results:
                if result["name"] == pelicula:
                    poster_path = result["poster_path"]
                    print(f"Portada de {pelicula} encontrada.")

                    # Construir la URL completa de la imagen de portada
                    poster_url = f"https://image.tmdb.org/t/p/w500/{poster_path}"

                    # Descargar la imagen de portada
                    poster_response = requests.get(poster_url)

                    # Guardar la imagen de portada en un archivo
                    file_name = re.sub(r'[<>:"/\\|?*\n]+', '_', pelicula).replace('.', '').replace(' ', '_').replace(':', '').replace('\'', '') + ".jpg"
                    with open(file_name, "wb") as f:
                        f.write(poster_response.content)
                    break
        else:
            print(f"No se encontró una portada para {pelicula}.")
    else:
        print(f"Tipo de película desconocido para {pelicula}.")